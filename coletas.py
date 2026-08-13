import os
import re
from io import BytesIO
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
import pandas as pd
import psycopg2
from fpdf import FPDF
from unidecode import unidecode
import openpyxl
import traceback
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import webbrowser
from threading import Timer
from dotenv import load_dotenv

load_dotenv()

# ================================================================
# CONFIGURAÇÕES
# ================================================================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.abspath(".")

DATABASE_URL = os.getenv("DATABASE_URL")

app = Flask(__name__)
app.secret_key = "chave_mestra_123"

DF_MATCH    = None
DF_FILTERED = None


# ================================================================
# CONEXÃO COM NEON
# ================================================================
def get_conn():
    return psycopg2.connect(DATABASE_URL)


def get_base_count():
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM base_coletas;")
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return count
    except Exception as e:
        print(f"[DB] Erro ao contar registros: {e}")
        return 0


# ================================================================
# FUNÇÕES DE NORMALIZAÇÃO
# ================================================================
def norm_text(s):
    s = unidecode(str(s)).upper()
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s.strip())


def norm_cep(s):
    s = re.sub(r'\D', '', str(s))
    return s.zfill(8) if s else ''


def clean_cell(s):
    """Remove o formato ="valor" gerado por exportações do Excel."""
    s = s.strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1]
    elif s.startswith('"') and s.endswith('"'):
        s = s[1:-1]
    return s.strip()


# ================================================================
# PARSING DO ARQUIVO DO DIA
# ================================================================
def try_decode_bytes(b: bytes):
    for enc in ('utf-8', 'latin-1', 'cp1252'):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode('utf-8', errors='ignore')


def parse_txt_to_df(path_or_bytes, is_bytes=False):
    """
    Detecta automaticamente o formato do arquivo e faz o parse correto.

    FORMATO A — linha 0 contém título + cabeçalho + 1º registro tudo junto (>=14 tabs).
    FORMATO B — título em linhas separadas; cabeçalho em linha própria com 14 tabs.

    Chave: Remetente|CEP — compatível com endereços completos ou truncados.
    """
    if is_bytes:
        text = try_decode_bytes(path_or_bytes)
    else:
        with open(path_or_bytes, 'rb') as f:
            text = try_decode_bytes(f.read())

    todas = text.splitlines()
    if not todas:
        return pd.DataFrame()

    linha0_tabs = todas[0].count('\t')

    if linha0_tabs >= 14:
        # ----------------------------------------------------------
        # FORMATO A: tudo na linha 0
        # ----------------------------------------------------------
        linhas = [l.strip() for l in todas if l.strip()]
        linha0_parts = linhas[0].split('\t')
        linha1_parts = linhas[1].split('\t') if len(linhas) > 1 else []
        n_cols = len(linha1_parts)

        header = linha0_parts[:n_cols]
        header[0] = 'Coleta'
        match_num = re.search(r'(\d{7,})', header[-1])
        first_pedido_num = match_num.group(1) if match_num else ''
        header[-1] = re.sub(r'\s+\d+.*$', '', header[-1]).strip()

        first_row_extra = linha0_parts[n_cols:]
        first_row = [first_pedido_num] + first_row_extra \
            if first_pedido_num and len(first_row_extra) == n_cols - 1 else None

        dados = []
        if first_row:
            dados.append([clean_cell(c) for c in first_row])
        for l in linhas[1:]:
            cols = l.split('\t')
            if len(cols) >= 5:
                if len(cols) < n_cols:
                    cols += [''] * (n_cols - len(cols))
                dados.append([clean_cell(c) for c in cols[:n_cols]])

        df = pd.DataFrame(dados, columns=header)

    else:
        # ----------------------------------------------------------
        # FORMATO B: cabeçalho em linha própria (14 tabs)
        # ----------------------------------------------------------
        cab_idx = None
        for i, l in enumerate(todas):
            if l.count('\t') >= 14 and 'coleta' in unidecode(l.lower()):
                cab_idx = i
                break

        if cab_idx is None:
            return pd.DataFrame()

        header = todas[cab_idx].split('\t')
        n_cols = len(header)

        dados = []
        for l in todas[cab_idx + 1:]:
            if not l.strip():
                continue
            cols = l.split('\t')
            if len(cols) >= 5:
                if len(cols) < n_cols:
                    cols += [''] * (n_cols - len(cols))
                dados.append([clean_cell(c) for c in cols[:n_cols]])

        df = pd.DataFrame(dados, columns=header)

    # ------------------------------------------------------------------
    # Mapeamento de colunas
    # ------------------------------------------------------------------
    colmap = {}
    for c in df.columns:
        c_low = unidecode(c.lower())
        if   'remetent'      in c_low: colmap[c] = 'Remetente'
        elif 'destinat'      in c_low: colmap[c] = 'Destinatario'
        elif 'endereco orig' in c_low: colmap[c] = 'EnderecoOrigem'
        elif 'cep orig'      in c_low: colmap[c] = 'CEPOrigem'
        elif 'status coleta' in c_low: colmap[c] = 'StatusColeta'

    df = df.rename(columns=colmap)

    for need in ('Remetente', 'EnderecoOrigem', 'CEPOrigem', 'Destinatario', 'StatusColeta'):
        if need not in df.columns:
            df[need] = ''

    # Chave: Remetente + CEP (sem endereço — compatível com todos os formatos)
    df['chave'] = df['Remetente'].map(norm_text) + '|' + df['CEPOrigem'].map(norm_cep)
    return df


# ================================================================
# CONSULTA NO BANCO
# ================================================================
def buscar_dados_por_chaves(chaves: list) -> dict:
    if not chaves:
        return {}
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute(
            """
            SELECT DISTINCT ON (chave) chave, numero_coleta, status_coleta
            FROM base_coletas
            WHERE chave = ANY(%s)
            """,
            (chaves,)
        )
        resultado = {row[0]: {'numero_coleta': row[1], 'status_coleta': row[2]} for row in cur.fetchall()}
        cur.close()
        conn.close()
        return resultado
    except Exception as e:
        print(f"[DB] Erro na consulta: {e}")
        return {}


# ================================================================
# HELPER — gera tabela HTML padronizada
# ================================================================
def df_to_html(df: pd.DataFrame) -> str:
    out = df[['NumeroColeta', 'CEPOrigem', 'Remetente', 'EnderecoOrigem', 'StatusColeta']].copy()
    out.columns = ['N° COLETA', 'CEP', 'REMETENTE', 'ENDEREÇO', 'STATUS (HISTÓRICO)']
    return out.to_html(
        classes='table table-sm table-striped table-bordered',
        index=False,
    )


# ================================================================
# ROTAS
# ================================================================

@app.route('/')
def index():
    base_count = get_base_count()
    return render_template('index.html', base_count=base_count)


@app.route('/upload_dia', methods=['POST'])
def upload_dia():
    global DF_MATCH, DF_FILTERED

    f = request.files.get('file')
    if not f:
        return redirect(url_for('index'))

    df_dia = parse_txt_to_df(f.read(), is_bytes=True)

    df_vivo = df_dia[df_dia['Destinatario'].str.contains("VIVO|TELEFON", case=False, na=False)].copy()
    total_vivo = len(df_vivo)

    chaves = df_vivo['chave'].unique().tolist()
    dados_map = buscar_dados_por_chaves(chaves)

    df_vivo['NumeroColeta'] = df_vivo['chave'].map(
        lambda c: dados_map[c]['numero_coleta'] if c in dados_map else None
    )
    df_vivo['StatusColeta'] = df_vivo['chave'].map(
        lambda c: dados_map[c]['status_coleta'] if c in dados_map else None
    )
    df_final = df_vivo[df_vivo['StatusColeta'].notna()].copy()

    if not df_final.empty:
        df_final['CEP_SORT'] = df_final['CEPOrigem'].str.replace(r'\D', '', regex=True)
        df_final = df_final.sort_values(by='CEP_SORT').drop(columns=['CEP_SORT'])

    DF_MATCH    = df_final.copy()
    DF_FILTERED = df_final.copy()

    tabela_html = df_to_html(DF_FILTERED)

    return render_template(
        'resultado.html',
        table=tabela_html,
        total_vivo=total_vivo,
        total_repetidos=len(DF_FILTERED),
        cep_min=None,
        cep_max=None,
    )


@app.route('/filtrar_cep', methods=['POST'])
def filtrar_cep():
    global DF_MATCH, DF_FILTERED

    if DF_MATCH is None or DF_MATCH.empty:
        return redirect(url_for('index'))

    cep_min_raw = request.form.get('cep_min', '').strip()
    cep_max_raw = request.form.get('cep_max', '').strip()

    cep_min_num = norm_cep(cep_min_raw)
    cep_max_num = norm_cep(cep_max_raw)

    if not cep_min_num or not cep_max_num:
        flash("Informe os dois CEPs para filtrar.", "warning")
        return redirect(url_for('index'))

    df = DF_MATCH.copy()
    df['_cep_num'] = df['CEPOrigem'].apply(norm_cep)
    DF_FILTERED = df[
        (df['_cep_num'] >= cep_min_num) & (df['_cep_num'] <= cep_max_num)
    ].drop(columns=['_cep_num']).copy()

    tabela_html = df_to_html(DF_FILTERED)

    def fmt_cep(c):
        c = re.sub(r'\D', '', c).zfill(8)
        return f"{c[:5]}-{c[5:]}" if len(c) == 8 else c

    return render_template(
        'resultado.html',
        table=tabela_html,
        total_vivo=len(DF_MATCH),
        total_repetidos=len(DF_FILTERED),
        cep_min=fmt_cep(cep_min_raw),
        cep_max=fmt_cep(cep_max_raw),
    )


@app.route('/limpar_filtro')
def limpar_filtro():
    global DF_MATCH, DF_FILTERED

    if DF_MATCH is None or DF_MATCH.empty:
        return redirect(url_for('index'))

    DF_FILTERED = DF_MATCH.copy()
    tabela_html = df_to_html(DF_FILTERED)

    return render_template(
        'resultado.html',
        table=tabela_html,
        total_vivo=len(DF_MATCH),
        total_repetidos=len(DF_FILTERED),
        cep_min=None,
        cep_max=None,
    )


@app.route('/download_pdf')
def download_pdf():
    global DF_FILTERED
    if DF_FILTERED is None or DF_FILTERED.empty:
        return redirect(url_for('index'))

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(190, 10, "Relatorio de Coletas Repetidas", ln=True, align='C')
    pdf.ln(5)

    pdf.set_font("Arial", 'B', 8)
    pdf.cell(28, 7, "N  COLETA", 1)
    pdf.cell(25, 7, "CEP", 1)
    pdf.cell(45, 7, "REMETENTE", 1)
    pdf.cell(60, 7, "ENDERECO", 1)
    pdf.cell(32, 7, "STATUS", 1, 1)

    pdf.set_font("Arial", '', 7)
    for _, row in DF_FILTERED.iterrows():
        pdf.cell(28, 6, str(row.get('NumeroColeta', ''))[:18], 1)
        pdf.cell(25, 6, str(row['CEPOrigem']), 1)
        pdf.cell(45, 6, unidecode(str(row['Remetente']))[:30], 1)
        pdf.cell(60, 6, unidecode(str(row['EnderecoOrigem']))[:45], 1)
        pdf.cell(32, 6, unidecode(str(row['StatusColeta']))[:22], 1)
        pdf.ln()

    out = BytesIO()
    pdf_str = pdf.output(dest='S').encode('latin-1', errors='ignore')
    out.write(pdf_str)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="repetidos.pdf")


@app.route('/relatorio_excel')
def relatorio_excel():
    """
    Gera Excel com coletas por cliente (histórico completo do banco).
    A agregação é feita no banco — só o resumo chega ao Python,
    evitando out-of-memory com bases grandes (670k+ registros).
    """
    SQL = """
        SELECT
            remetente,
            cep_origem,
            COUNT(*)                                                                AS total_coletas,
            COUNT(*) FILTER (WHERE TRIM(status_coleta) ILIKE 'coleta cancelada')   AS coleta_cancelada,
            COUNT(*) FILTER (WHERE TRIM(status_coleta) ILIKE '%tentativa%')        AS segunda_tentativa
        FROM base_coletas
        WHERE destinatario ILIKE '%vivo%' OR destinatario ILIKE '%telefonica%'
        GROUP BY remetente, cep_origem
        HAVING COUNT(*) > 1
        ORDER BY total_coletas DESC
    """
    try:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute(SQL)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[RELATORIO] Erro SQL: {e}", flush=True)
        flash(f"Erro ao consultar banco: {e}", "danger")
        return redirect(url_for('index'))

    try:
        pass  # validação de rows
        print(f"[RELATORIO] {len(rows)} clientes encontrados", flush=True)
    except Exception as e:
        print(f"[RELATORIO] Erro inesperado: {e}", flush=True)
        flash(f"Erro inesperado: {e}", "danger")
        return redirect(url_for('index'))

    try:
        _excel_out = _gerar_excel_relatorio(rows)
        nome = f"relatorio_coletas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            _excel_out,
            as_attachment=True,
            download_name=nome,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"[RELATORIO] Erro ao gerar Excel: {e}\n{traceback.format_exc()}", flush=True)
        flash(f"Erro ao gerar Excel: {e}", "danger")
        return redirect(url_for('index'))


def _gerar_excel_relatorio(rows):
    NAVY   = "0A1628"
    ACCENT = "2563EB"
    WHITE  = "FFFFFF"
    GRAY   = "F1F5F9"
    RED    = "DC2626"
    BC     = "DDE5F5"

    # Criar objetos de borda UMA ÚNICA VEZ — reutilizar no loop
    # (recriar Border() por célula causa recursão infinita no openpyxl/Python 3.14)
    _side_thin   = Side(style="thin",   color=BC)
    _side_medium = Side(style="medium", color=NAVY)
    BORDER_CELL  = Border(
        left=_side_thin, right=_side_thin,
        top=_side_thin,  bottom=_side_thin
    )
    BORDER_TOTAL = Border(
        left=_side_thin,   right=_side_thin,
        top=_side_medium,  bottom=_side_medium
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coletas por Cliente"

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = "Relatório de Coletas Repetidas — VIVO / Telefônica"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"] = (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f"  •  Histórico completo  •  {len(rows)} clientes")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="64748B")
    ws["A2"].fill      = PatternFill("solid", fgColor="E8F0FE")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Cabeçalho
    HEADERS = ["REMETENTE", "CEP", "TOTAL COLETAS", "COLETA CANCELADA", "2ª TENTATIVA / CANCELADA"]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER_CELL
    ws.row_dimensions[3].height = 28

    # Pré-criar fills para não recriar por linha
    FILL_ODD  = PatternFill("solid", fgColor=WHITE)
    FILL_EVEN = PatternFill("solid", fgColor=GRAY)

    # Dados
    for i, (remetente, cep, total, cancelada, tentativa) in enumerate(rows, 1):
        rn   = i + 3
        fill = FILL_EVEN if i % 2 == 0 else FILL_ODD

        for col, val, align, bold, color in [
            (1, remetente, "left",   False, NAVY),
            (2, cep,       "center", False, NAVY),
            (3, total,     "center", True,  NAVY),
            (4, cancelada, "center", False, RED if cancelada > 0 else NAVY),
            (5, tentativa, "center", False, RED if tentativa > 0 else NAVY),
        ]:
            c = ws.cell(row=rn, column=col, value=val)
            c.font      = Font(name="Arial", size=9, bold=bold, color=color)
            c.fill      = fill
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = BORDER_CELL

    # Linha de total
    tr = len(rows) + 4
    ws.merge_cells(f"A{tr}:B{tr}")
    for col, val in [
        (1, "TOTAL GERAL"),
        (3, f"=SUM(C4:C{tr-1})"),
        (4, f"=SUM(D4:D{tr-1})"),
        (5, f"=SUM(E4:E{tr-1})")
    ]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER_TOTAL
    ws.row_dimensions[tr].height = 24

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "A4"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


if __name__ == "__main__":
    Timer(1, lambda: webbrowser.open("http://localhost:5000")).start()
    app.run(port=5000, debug=False)
